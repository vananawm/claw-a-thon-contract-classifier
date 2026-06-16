#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py — Web app upload cho Agent phân loại hợp đồng (MVP cho AgentBase).

Tính năng:
  - 2 khu upload riêng: Template/hợp đồng mẫu  &  Hợp đồng cần phân loại
  - Tùy chọn giữ lại template đã upload để dùng cho các lần so sánh sau
  - Kết quả: danh sách tên file theo 3 nhóm (>=70% / trọng yếu / không theo template)
  - Tải báo cáo Excel và tải gói kết quả (zip) đã sắp theo nhóm

Chạy local:   python app.py   ->   http://localhost:8080
"""

import os, io, re, zipfile, datetime, tempfile, shutil
try:                                  # nạp .env khi chạy local (trên cloud env có sẵn)
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass
from flask import Flask, request, render_template_string, send_file, redirect, url_for
import classifier as C

BASE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_STORE = os.path.join(BASE, "template_store")   # template giữ lại giữa các lần
RUNS_DIR       = os.path.join(BASE, "web_runs")          # kết quả mỗi lần chạy
os.makedirs(TEMPLATE_STORE, exist_ok=True)
os.makedirs(RUNS_DIR, exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 80 * 1024 * 1024  # 80MB

# Màu nhãn (pill) và màu tiêu đề nhóm theo từng category
PILL = {C.CAT_SIM: "p0", C.CAT_NONMATERIAL: "p0", C.CAT_KEY: "p1", C.CAT_NONE: "p2"}
GRP  = {C.CAT_SIM: "g0", C.CAT_NONMATERIAL: "g0", C.CAT_KEY: "g1", C.CAT_NONE: "g2"}

def safe_name(filename):
    """Giữ tên file (kể cả tiếng Việt), chỉ loại bỏ ký tự đường dẫn."""
    name = os.path.basename(filename or "").replace("\\", "_").replace("/", "_").strip()
    return name or "file"

def stored_templates():
    return sorted(f for f in os.listdir(TEMPLATE_STORE)
                  if os.path.splitext(f)[1].lower() in C.SUPPORTED_EXT)

# ----------------- Giao diện -----------------
PAGE = """
<!doctype html><html lang="vi"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Agent Phân Loại Hợp Đồng</title>
<style>
 body{font-family:system-ui,Segoe UI,Arial,sans-serif;background:#f5f7f6;color:#1f2937;margin:0}
 .wrap{max-width:880px;margin:0 auto;padding:28px 20px 60px}
 h1{color:#0b7a3b;margin:0 0 4px}
 .sub{color:#6b7280;margin:0 0 24px}
 .card{background:#fff;border:1px solid #e5e7eb;border-radius:14px;padding:20px;margin-bottom:18px;box-shadow:0 1px 2px rgba(0,0,0,.04)}
 .card h2{margin:0 0 6px;font-size:18px}
 .hint{color:#6b7280;font-size:13px;margin:0 0 12px}
 input[type=file]{display:block;margin:8px 0;font-size:14px}
 .row{display:flex;align-items:center;gap:8px;margin-top:10px;font-size:14px}
 .btn{background:#0b7a3b;color:#fff;border:0;border-radius:10px;padding:12px 22px;font-size:15px;cursor:pointer}
 .btn:hover{background:#0a6c35}
 .chips span{display:inline-block;background:#eef2ff;color:#3730a3;border-radius:20px;padding:4px 10px;margin:3px 4px 0 0;font-size:12px}
 table{width:100%;border-collapse:collapse;margin-top:8px;font-size:14px}
 th,td{border:1px solid #e5e7eb;padding:8px 10px;text-align:left}
 th{background:#0b7a3b;color:#fff}
 .grp{font-weight:700;margin:18px 0 4px;font-size:15px}
 .g0{color:#0b7a3b}.g1{color:#b45309}.g2{color:#6b7280}
 .dl a{display:inline-block;margin:6px 10px 0 0;background:#111827;color:#fff;text-decoration:none;padding:9px 16px;border-radius:9px;font-size:14px}
 .pill{font-size:12px;border-radius:20px;padding:2px 9px}
 .p0{background:#dcfce7;color:#166534}.p1{background:#fef3c7;color:#92400e}.p2{background:#f3f4f6;color:#374151}
 .head{display:flex;align-items:center;gap:14px;margin-bottom:4px}
 .head h1{font-size:26px}
 .opt{font-size:12px;font-weight:600;color:#0b7a3b;background:#dcfce7;border-radius:20px;padding:2px 10px;margin-left:6px}
 .note-m{color:#b91c1c;font-weight:600}.note-n{color:#374151}
</style></head><body><div class="wrap">
<div class="head">
  <svg width="62" height="62" viewBox="0 0 72 72" xmlns="http://www.w3.org/2000/svg">
    <rect width="72" height="72" rx="16" fill="#0E9F43"/>
    <rect x="16" y="13" width="30" height="44" rx="5" fill="#ffffff" opacity="0.4" transform="rotate(-7 31 35)"/>
    <rect x="19" y="14" width="31" height="44" rx="5" fill="#ffffff"/>
    <rect x="24" y="20" width="14" height="5" rx="2.5" fill="#13B24B"/>
    <rect x="24" y="29" width="21" height="3.4" rx="1.7" fill="#D7DEDA"/>
    <rect x="24" y="36" width="21" height="3.4" rx="1.7" fill="#D7DEDA"/>
    <rect x="24" y="43" width="15" height="3.4" rx="1.7" fill="#D7DEDA"/>
    <rect x="46" y="24" width="12" height="7" rx="2" fill="#16A34A"/>
    <rect x="46" y="33" width="12" height="7" rx="2" fill="#F59E0B"/>
    <circle cx="50" cy="50" r="11" fill="#0E9F43" stroke="#ffffff" stroke-width="2.5"/>
    <path d="M45 50 l3.5 3.5 6.5 -7" fill="none" stroke="#ffffff" stroke-width="2.6" stroke-linecap="round" stroke-linejoin="round"/>
  </svg>
  <h1>Agent Phân Loại Hợp Đồng</h1>
</div>
<p class="sub">Đối chiếu hợp đồng với template &amp; nhận diện hợp đồng trọng yếu bằng model Qwen. Có thể chạy không cần template.</p>

{% if not result %}
<form method="post" action="{{ url_for('classify') }}" enctype="multipart/form-data">
  <div class="card">
    <h2>1) Template / Hợp đồng mẫu <span class="opt">Tùy chọn</span></h2>
    <p class="hint">Bộ template chuẩn để đối chiếu (.docx, .doc, .pdf). <b>Không bắt buộc</b> — nếu bỏ trống, agent chỉ phân loại trọng yếu / không trọng yếu.</p>
    <input type="file" name="templates" multiple accept=".docx,.doc,.pdf,.txt,.md">
    {% if stored %}
      <div class="row"><input type="checkbox" name="use_stored" id="us" checked>
        <label for="us">Dùng cả {{ stored|length }} template đã lưu trước đó:</label></div>
      <div class="chips">{% for s in stored %}<span>{{ s }}</span>{% endfor %}</div>
    {% endif %}
    <div class="row"><input type="checkbox" name="keep_templates" id="kt">
      <label for="kt">Giữ lại template vừa upload để dùng cho các lần so sánh sau</label></div>
  </div>
  <div class="card">
    <h2>2) Hợp đồng cần phân loại</h2>
    <p class="hint">Các hợp đồng cần kiểm tra (.docx, .doc, .pdf). Có thể upload nhiều file.</p>
    <input type="file" name="contracts" multiple required accept=".docx,.doc,.pdf,.txt,.md">
  </div>
  <button class="btn" type="submit">Phân tích</button>
</form>
{% else %}
  <div class="card">
    <h2>Kết quả phân loại</h2>
    <p class="hint">{{ results|length }} hợp đồng · {{ n_templates }} template đối chiếu{% if not has_templates %} (không dùng template){% endif %} · {{ run_id }}</p>
    {% for cat in cats %}
      <div class="grp {{ grp[cat] }}">{{ cat }} ({{ by_cat[cat]|length }})</div>
      {% if by_cat[cat] %}
        <div class="chips">{% for n in by_cat[cat] %}<span>{{ n }}</span>{% endfor %}</div>
      {% else %}<p class="hint">(trống)</p>{% endif %}
    {% endfor %}
    <table><tr><th>Hợp đồng</th>{% if has_templates %}<th>Template giống nhất</th><th>% giống</th>{% endif %}<th>Nhóm</th><th>Loại (AI)</th><th>Quy trình review</th></tr>
    {% for r in results %}<tr>
      <td>{{ r.name }}</td>
      {% if has_templates %}<td>{{ r.best_tpl }}</td><td>{{ r.sim }}%</td>{% endif %}
      <td><span class="pill {{ pill[r.cat] }}">{{ r.cat }}</span></td>
      <td>{% if r.loai %}<b>{{ r.loai }}</b>{% endif %}{% if r.ly_do %}<br><span class="hint">{{ r.ly_do }}</span>{% endif %}</td>
      <td class="{{ 'note-m' if r.material else 'note-n' }}">{{ r.note }}</td>
    </tr>{% endfor %}</table>
    <div class="dl">
      <a href="{{ url_for('download', run_id=run_id, fname='report.xlsx') }}">⬇ Tải báo cáo Excel</a>
      <a href="{{ url_for('download', run_id=run_id, fname='ket_qua_phan_loai.zip') }}">⬇ Tải gói kết quả (zip)</a>
      <a href="{{ url_for('home') }}" style="background:#0b7a3b">↺ Phân tích lô khác</a>
    </div>
  </div>
{% endif %}
<p class="sub" style="margin-top:24px;font-size:12px">⚠️ Chỉ dùng hợp đồng công khai / giả lập / đã ẩn danh — không dùng dữ liệu thật của khách hàng hay nội bộ.</p>
</div></body></html>
"""

@app.route("/health")
def health():
    # AgentBase Runtime gọi endpoint này để đánh dấu agent ACTIVE
    return "ok", 200

@app.route("/")
def home():
    return render_template_string(PAGE, result=False, stored=stored_templates())

@app.route("/classify", methods=["POST"])
def classify():
    # 1) Thu thập template
    run_id = "R-" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = os.path.join(RUNS_DIR, run_id)
    tdir = os.path.join(run_dir, "_templates")
    os.makedirs(tdir, exist_ok=True)
    template_paths = []

    if request.form.get("use_stored"):
        for s in stored_templates():
            template_paths.append(os.path.join(TEMPLATE_STORE, s))

    keep = bool(request.form.get("keep_templates"))
    for f in request.files.getlist("templates"):
        if not f or not f.filename: continue
        p = os.path.join(tdir, safe_name(f.filename)); f.save(p)
        template_paths.append(p)
        if keep:
            shutil.copy(p, os.path.join(TEMPLATE_STORE, safe_name(f.filename)))

    templates = C.build_templates(template_paths)
    has_templates = bool(templates)
    cats = C.CATEGORIES_WITH_TEMPLATE if has_templates else C.CATEGORIES_NO_TEMPLATE

    # 2) Lưu & phân loại hợp đồng (template TÙY CHỌN)
    cdir = os.path.join(run_dir, "_contracts"); os.makedirs(cdir, exist_ok=True)
    results = []
    for f in request.files.getlist("contracts"):
        if not f or not f.filename: continue
        name = safe_name(f.filename)
        p = os.path.join(cdir, name); f.save(p)
        results.append((p, C.classify_file(p, templates, display_name=name)))

    # 3) Sắp file vào folder theo nhóm + report + zip
    out_dir = os.path.join(run_dir, "ket_qua")
    for c in cats: os.makedirs(os.path.join(out_dir, c), exist_ok=True)
    for path, r in results:
        shutil.copy(path, os.path.join(out_dir, r["cat"], r["name"]))
    res_list = [r for _, r in results]
    write_report_xlsx(os.path.join(run_dir, "report.xlsx"), res_list, cats)
    make_zip(os.path.join(run_dir, "ket_qua_phan_loai.zip"), out_dir,
             os.path.join(run_dir, "report.xlsx"))

    by_cat = {c: [r["name"] for r in res_list if r["cat"] == c] for c in cats}
    return render_template_string(PAGE, result=True, results=res_list, by_cat=by_cat,
        cats=cats, grp=GRP, pill=PILL, has_templates=has_templates,
        n_templates=len(templates), run_id=run_id, stored=stored_templates())

@app.route("/download/<run_id>/<path:fname>")
def download(run_id, fname):
    safe = re.sub(r"[^A-Za-z0-9_.\-]", "", run_id)
    p = os.path.join(RUNS_DIR, safe, os.path.basename(fname))
    if not os.path.exists(p): return "Không tìm thấy file", 404
    return send_file(p, as_attachment=True)

def write_report_xlsx(path, results, cats):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tóm tắt"
        by_cat = {c: [r["name"] for r in results if r["cat"] == c] for c in cats}
        ws.append(["Nhóm", "Số lượng", "Các hợp đồng"])
        for c in cats: ws.append([c, len(by_cat[c]), ", ".join(by_cat[c])])
        ws2 = wb.create_sheet("Chi tiết")
        ws2.append(["Hợp đồng", "Tiêu đề", "Template giống nhất", "% giống", "Nhóm",
                    "Loại (AI)", "Lý do (AI)", "Quy trình review"])
        for r in results:
            ws2.append([r["name"], r["title"], r["best_tpl"], r["sim"], r["cat"],
                        r.get("loai", ""), r.get("ly_do", ""), r.get("note", "")])
        for w in (ws, ws2):
            for cell in w[1]:
                cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D9EAD3")
        wb.save(path)
    except Exception as e:
        print("xlsx error:", e)

def make_zip(zip_path, out_dir, report_path):
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(out_dir):
            for fn in files:
                fp = os.path.join(root, fn)
                z.write(fp, os.path.relpath(fp, os.path.dirname(out_dir)))
        if os.path.exists(report_path):
            z.write(report_path, "report.xlsx")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8080"))
    app.run(host="0.0.0.0", port=port, debug=False)
