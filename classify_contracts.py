#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
classify_contracts.py — Bản chạy theo FOLDER (chế độ batch / local).

Đọc các folder hợp đồng LG-* trong contracts/, đối chiếu với toàn bộ template
trong template/, phân vào 3 nhóm, copy folder LG-* vào Result/R-<timestamp>/
và xuất báo cáo ngắn.

Lõi phân loại nằm ở classifier.py (dùng chung với web app).
"""

import os, re, glob, shutil, datetime
import classifier as C

BASE = os.path.dirname(os.path.abspath(__file__))
CONTRACTS_DIR = os.path.join(BASE, "contracts")
TEMPLATE_DIR  = os.path.join(BASE, "template")
RESULT_DIR    = os.path.join(BASE, "Result")

def load_contract_folder(folder):
    """Gộp toàn bộ file trong folder LG-* thành nội dung 1 hợp đồng."""
    texts = []
    for f in sorted(glob.glob(os.path.join(folder, "**", "*"), recursive=True)):
        if os.path.isfile(f) and os.path.splitext(f)[1].lower() in C.SUPPORTED_EXT:
            t = C.extract_text(f)
            if t.strip():
                texts.append(t)
    return "\n".join(texts)

def main():
    print("== Agent phân loại hợp đồng (chế độ folder) ==")
    print("[1/4] Nạp template...")
    tpl_paths = [f for f in sorted(glob.glob(os.path.join(TEMPLATE_DIR, "*")))
                 if os.path.isfile(f)]
    templates = C.build_templates(tpl_paths)
    for t in templates:
        print(f"   + {t['name']}")
    if not templates:
        print("Không có template nào đọc được. Dừng."); return

    print("[2/4] Đọc các folder hợp đồng LG-*...")
    lg_folders = sorted([d for d in glob.glob(os.path.join(CONTRACTS_DIR, "LG-*"))
                         if os.path.isdir(d)])
    if not lg_folders:
        print("Không tìm thấy folder LG-* trong contracts/. Dừng."); return

    results = []
    for folder in lg_folders:
        lg = os.path.basename(folder)
        text = load_contract_folder(folder)
        r = C.classify_text(lg, text, templates)
        results.append(r)
        print(f"   {lg}: {r['sim']}% ~ {r['best_tpl']}  ->  {r['cat']}")

    print("[3/4] Tạo Result/R-<timestamp> và copy folder...")
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = os.path.join(RESULT_DIR, f"R-{ts}")
    for c in C.CATEGORIES:
        os.makedirs(os.path.join(run_dir, c), exist_ok=True)
    for r in results:
        shutil.copytree(os.path.join(CONTRACTS_DIR, r["name"]),
                        os.path.join(run_dir, r["cat"], r["name"]),
                        dirs_exist_ok=True)

    print("[4/4] Xuất báo cáo...")
    write_reports(run_dir, ts, results, templates)
    print(f"\nXong! Kết quả ở: {run_dir}")

def write_reports(run_dir, ts, results, templates):
    by_cat = {c: [r["name"] for r in results if r["cat"] == c] for c in C.CATEGORIES}
    lines = [f"# Báo cáo phân loại hợp đồng — R-{ts}", "",
             f"Tổng số hợp đồng: {len(results)}  |  Số template đối chiếu: {len(templates)}", "",
             "## Tóm tắt: mỗi nhóm chứa các folder LG- nào"]
    for c in C.CATEGORIES:
        items = ", ".join(by_cat[c]) if by_cat[c] else "(trống)"
        lines.append(f"- **{c}** ({len(by_cat[c])}): {items}")
    lines += ["", "## Chi tiết",
              "| Folder | Tiêu đề | Template giống nhất | % giống | Nhóm |",
              "|---|---|---|---|---|"]
    for r in results:
        lines.append(f"| {r['name']} | {r['title']} | {r['best_tpl']} | {r['sim']}% | {r['cat']} |")
    md = "\n".join(lines)
    open(os.path.join(run_dir, "report.md"), "w", encoding="utf-8").write(md)
    open(os.path.join(run_dir, "report.txt"), "w", encoding="utf-8").write(re.sub(r"[#*|`]", "", md))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tóm tắt"
        ws.append(["Nhóm", "Số lượng", "Các folder LG-"])
        for c in C.CATEGORIES:
            ws.append([c, len(by_cat[c]), ", ".join(by_cat[c])])
        ws2 = wb.create_sheet("Chi tiết")
        ws2.append(["Folder", "Tiêu đề", "Template giống nhất", "% giống", "Nhóm"])
        for r in results:
            ws2.append([r["name"], r["title"], r["best_tpl"], r["sim"], r["cat"]])
        for w in (ws, ws2):
            for cell in w[1]:
                cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D9EAD3")
        wb.save(os.path.join(run_dir, "report.xlsx"))
    except Exception as e:
        print(f"   [bỏ qua xlsx] {e}")

if __name__ == "__main__":
    main()
